"""
fetch_ssci_journals.py - 基于 JCR SSCI/AHCI 期刊名单，从 OpenAlex 拉取完整元数据

策略变化：不再用 topic_share 过滤（覆盖不全且不稳定），
改为直接用 JCR 2026 的 ISSN 列表作为种子，从 OpenAlex 获取每个期刊的 topics 等详细数据。

采集范围（两个通道，任一命中即收）：
1. SSCI / AHCI —— 社科人文主库（原有逻辑）
2. 环境健康交叉 SCIE 白名单 —— 补足 climate×health×demography 的自然科学侧主场刊。
   - 公共卫生分类（PUBLIC, ENVIRONMENTAL & OCCUPATIONAL HEALTH）：净新增全收
   - 环境科学分类（ENVIRONMENTAL SCIENCES）：用名称黑名单剔除环境化工/生态/水处理
     等与环境健康无关的子领域，压低语义噪声

每条记录带 `_source_scope` 标记（ssci_ahci / scie_env_health），便于区分与调试。

输入：JCR Excel 中 SSCI/AHCI + 环境健康 SCIE 白名单期刊的 ISSN
输出：data/raw/sources_ssci_all.json
"""

import argparse
import json
import os
import time
import requests
import openpyxl
from pathlib import Path
from tqdm import tqdm

BASE_URL = "https://api.openalex.org"

# Load API key from .env file or environment variable
_env_file = Path(__file__).parent.parent / ".env"
if _env_file.exists():
    for line in _env_file.read_text().splitlines():
        if line.startswith("OPENALEX_KEY="):
            os.environ.setdefault("OPENALEX_KEY", line.split("=", 1)[1].strip())
        elif line.startswith("OPENALEX_MAILTO="):
            os.environ.setdefault("OPENALEX_MAILTO", line.split("=", 1)[1].strip())
API_KEY = os.environ.get("OPENALEX_KEY", "")
MAILTO = os.environ.get("OPENALEX_MAILTO", "")

JCR_FILE = Path.home() / "Downloads/期刊/2026年度JCR期刊名单（完整版）.xlsx"
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "raw"
OUTPUT_FILE = OUTPUT_DIR / "sources_ssci_all.json"

# 环境健康交叉 SCIE 白名单：匹配 JCR col[31] 分区详情里的分类名（大写）
# 公共卫生分类全收；环境科学分类需再过名称黑名单去化工。
SCIE_HEALTH_CATEGORY_ALL = "PUBLIC, ENVIRONMENTAL & OCCUPATIONAL HEALTH"
SCIE_HEALTH_CATEGORY_FILTERED = "ENVIRONMENTAL SCIENCES"

# 环境科学分类下需剔除的非健康子领域关键词（匹配期刊名，大写）。
# 目的：滤掉环境化工/生态/水处理/地学等与温度-健康研究无关的刊，压低语义噪声。
SCIE_ENV_NAME_BLACKLIST = [
    "WATER", "AQUATIC", "MARINE", "FRESHWATER", "OCEAN", "COAST", "HYDRO", "SOIL",
    "AGRICULT", "FORAGE", "RANGE", "BOTAN", "PLANT", "FOREST", "ECOLOG", "BIODIVERS",
    "CONSERVAT", "ECOSYSTEM", "WILDLIFE", "CATALY", "CHEMISTR", "CHEMICAL", "BIOTECHNOL",
    "BIOENERG", "BIOCHAR", "AEROSOL", "METEOROL", "REMOTE SENSING", "SPACE", "ARCTIC",
    "ANTARCTIC", "ALPINE", "BOREAL", "ANTHROPOCENE", "CIRCULAR", "WASTE", "CLEANER",
    "CLEAN-", "CLEAN TECHNOL", "ENERGY", "CARBON", "MUTAGEN", "FLUID", "FORENS",
    "ENGINEER", "ENVIRONMETRIC", "MODELLING", "MODELING", "SOFTWARE", "STATISTIC",
    "GEOGRAPH", "EARTH", "GEOSCI", "BOTANY", "REMOTE", "AEROBIOLOG", "TECHNOLOGY",
    "ECOHYDROL", "BIOGEOCHEM", "HAZARDOUS", "MICROPLASTIC", "NANO", "GEOBIOL",
    "GEOMICROBIOL", "PALEOLIMN", "SEDIMENT", "RADIOACT", "RADIATION", "MINING",
    "TRACE ELEMENT", "ELEMENTOLOGY", "LIMNOL", "PEAT", "WETLAND", "RIVER", "LAKE",
    "MICROB", "VIROL", "DEGRAD", "PHYTOR", "MOUNTAIN", "POLAR", "GLACIER", "VADOSE",
    "ARID", "ISOTOPE", "BIODETERIOR", "BIODEGRAD", "QSAR", "SAR AND", "PALEO",
    "GRUNDWASSER", "GREAT LAKES", "SPATIAL", "GEOVISU", "GEOCARTO", "NITROGEN",
    "MINERAL",
]


def classify_scope(name, detail):
    """判断一本期刊属于哪个采集通道，不在范围内返回 None。

    detail 为 JCR col[31]（各学科分区详情），已大写。
    返回 'ssci_ahci' 或 'scie_env_health'。
    """
    if "SSCI" in detail or "AHCI" in detail:
        return "ssci_ahci"

    # 公共卫生分类：净新增全收
    if SCIE_HEALTH_CATEGORY_ALL in detail:
        return "scie_env_health"

    # 环境科学分类：过名称黑名单去化工
    if SCIE_HEALTH_CATEGORY_FILTERED in detail:
        name_upper = (name or "").upper()
        if not any(b in name_upper for b in SCIE_ENV_NAME_BLACKLIST):
            return "scie_env_health"

    return None


def find_jcr_file(default_path):
    """Locate the JCR Excel even if Downloads was reorganized."""
    if default_path.exists():
        return default_path
    for root in [Path.home() / "Downloads", Path.home() / "Desktop", Path.home() / "Documents"]:
        if not root.exists():
            continue
        for pattern in ["*JCR*.xlsx", "*jcr*.xlsx", "*期刊名单*.xlsx"]:
            matches = sorted(root.rglob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
            if matches:
                return matches[0]
    return default_path


def create_session():
    session = requests.Session()
    contact = f" (mailto:{MAILTO})" if MAILTO else ""
    headers = {
        "User-Agent": f"JournalFinder/1.0{contact}",
    }
    if API_KEY:
        headers["Authorization"] = f"Bearer {API_KEY}"
        print("  Using API key from environment")
    else:
        print("  No API key found, using polite pool only")
    session.headers.update(headers)
    return session


def save_json_atomic(path, data):
    """Write JSON without exposing a partially written canonical seed."""
    temp_path = path.with_suffix(path.suffix + ".tmp")
    with open(temp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    temp_path.replace(path)


def load_ssci_issns():
    """从 JCR Excel 中提取所有 SSCI/AHCI 期刊的 ISSN。"""
    print("Loading JCR 2026 Excel...")
    wb = openpyxl.load_workbook(JCR_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]

    journals = []
    scope_counts = {"ssci_ahci": 0, "scie_env_health": 0}
    # 数据从第 2 行开始（第 1 行是表头）；detail 在 col[31]。
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1] or not row[31]:
            continue
        name = str(row[1]).strip()
        detail = str(row[31]).upper()

        scope = classify_scope(name, detail)
        if scope is None:
            continue

        issn = str(row[3]).strip() if row[3] else None
        eissn = str(row[4]).strip() if row[4] else None
        subject = str(row[6]).strip() if row[6] else ""

        journals.append({
            "name": name,
            "issn": issn,
            "eissn": eissn,
            "subject": subject,
            "jcr_detail": str(row[31]).strip() if row[31] else "",
            "scope": scope,
        })
        scope_counts[scope] += 1

    wb.close()
    print(
        f"  Found {len(journals)} journals in JCR "
        f"(SSCI/AHCI: {scope_counts['ssci_ahci']}, "
        f"SCIE env-health: {scope_counts['scie_env_health']})"
    )
    return journals


def normalize_issn(value):
    """Normalize print/eISSN values for reliable cross-file matching."""
    if not value:
        return None
    compact = "".join(ch for ch in str(value).upper() if ch.isdigit() or ch == "X")
    if len(compact) == 8:
        return f"{compact[:4]}-{compact[4:]}"
    return str(value).strip().upper()


def reconcile_existing_records(existing_data, jcr_journals):
    """Use the current JCR whitelist as the canonical seed for resume runs.

    Existing OpenAlex records are reused when either the JCR print ISSN or
    eISSN matches, but their JCR-derived fields are always refreshed. Records
    absent from the current whitelist are omitted from the returned results.
    """
    existing_by_issn = {}
    for index, journal in enumerate(existing_data):
        issns = journal.get("issn", [])
        if isinstance(issns, str):
            issns = [issns]
        for value in [journal.get("issn_l"), *issns]:
            normalized = normalize_issn(value)
            if normalized:
                existing_by_issn.setdefault(normalized, index)

    retained = []
    missing = []
    used_existing = set()
    refreshed = 0

    for jcr in jcr_journals:
        match_index = None
        for value in [jcr.get("issn"), jcr.get("eissn")]:
            normalized = normalize_issn(value)
            if normalized in existing_by_issn:
                match_index = existing_by_issn[normalized]
                break

        if match_index is None:
            missing.append(jcr)
            continue
        if match_index in used_existing:
            continue

        journal = existing_data[match_index].copy()
        refreshed_fields = {
            "_jcr_subject": jcr.get("subject", ""),
            "_jcr_detail": jcr.get("jcr_detail", ""),
            "_source_scope": jcr.get("scope", ""),
        }
        if any(journal.get(key) != value for key, value in refreshed_fields.items()):
            refreshed += 1
        journal.update(refreshed_fields)

        retained.append(journal)
        used_existing.add(match_index)

    stats = {
        "retained": len(retained),
        "refreshed": refreshed,
        "pruned": len(existing_data) - len(used_existing),
        "missing": len(missing),
    }
    return retained, missing, stats


def fetch_source_by_issn(session, issn):
    """Query OpenAlex for a single source by ISSN."""
    mailto_param = f"&mailto={MAILTO}" if MAILTO else ""
    url = f"{BASE_URL}/sources?filter=issn:{issn}{mailto_param}"
    for attempt in range(3):
        try:
            resp = session.get(url, timeout=30)
            if resp.status_code == 200:
                data = resp.json()
                if data.get("results"):
                    return data["results"][0]
                return None
            if resp.status_code == 429:
                # Rate limited - wait and retry
                time.sleep(10)
                continue
            if resp.status_code >= 500:
                time.sleep(2)
                continue
        except requests.RequestException:
            time.sleep(2)
    return None


def extract_journal_data(source, jcr_info):
    """Extract relevant fields from an OpenAlex source record."""
    summary = source.get("summary_stats", {})
    topics = source.get("topics", [])

    return {
        "openalex_id": source.get("id", "").split("/")[-1],
        "issn_l": source.get("issn_l"),
        "issn": source.get("issn", []),
        "name": source.get("display_name"),
        "alternate_titles": source.get("alternate_titles", []),
        "publisher": source.get("host_organization_name"),
        "country_code": source.get("country_code"),
        "homepage_url": source.get("homepage_url"),
        "is_oa": source.get("is_oa", False),
        "is_in_doaj": source.get("is_in_doaj", False),
        "apc_usd": source.get("apc_usd"),
        "apc_prices": source.get("apc_prices", []),
        "works_count": source.get("works_count", 0),
        "cited_by_count": source.get("cited_by_count", 0),
        "citedness_2yr": summary.get("2yr_mean_citedness", 0),
        "h_index": summary.get("h_index", 0),
        "topics": [
            {
                "id": t["id"].split("/")[-1],
                "name": t["display_name"],
                "count": t.get("count", 0),
                "subfield": t.get("subfield", {}).get("display_name", ""),
                "field": t.get("field", {}).get("display_name", ""),
            }
            for t in topics[:15]
        ],
        "type": source.get("type"),
        # Carry JCR metadata for later use
        "_jcr_subject": jcr_info.get("subject", ""),
        "_jcr_detail": jcr_info.get("jcr_detail", ""),
        "_source_scope": jcr_info.get("scope", ""),
    }


def main():
    global JCR_FILE

    parser = argparse.ArgumentParser(
        description="Fetch canonical journal metadata from a JCR Excel whitelist"
    )
    parser.add_argument(
        "--jcr-file",
        type=Path,
        help="JCR Excel file; otherwise search common user document directories",
    )
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    JCR_FILE = (
        args.jcr_file.expanduser().resolve()
        if args.jcr_file
        else find_jcr_file(JCR_FILE)
    )
    print(f"JCR source: {JCR_FILE}")
    session = create_session()

    # Load SSCI ISSNs from JCR
    jcr_journals = load_ssci_issns()

    # Reconcile existing results against the current canonical JCR whitelist.
    existing_data = []
    if OUTPUT_FILE.exists():
        with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
            existing_data = json.load(f)
        print(f"  Found {len(existing_data)} existing records (resuming)")

    results, remaining, resume_stats = reconcile_existing_records(
        existing_data,
        jcr_journals,
    )
    print(
        "  Resume reconciliation: "
        f"retained={resume_stats['retained']}, "
        f"refreshed={resume_stats['refreshed']}, "
        f"pruned={resume_stats['pruned']}, "
        f"missing={resume_stats['missing']}"
    )
    print(f"  Remaining to fetch: {len(remaining)}")

    not_found = 0
    fetched_this_run = 0
    for i, jcr in enumerate(tqdm(remaining, desc="Fetching from OpenAlex")):
        # Try ISSN first, then eISSN
        source = None
        if jcr.get("issn"):
            source = fetch_source_by_issn(session, jcr["issn"])
        if not source and jcr.get("eissn"):
            source = fetch_source_by_issn(session, jcr["eissn"])

        if source:
            journal_data = extract_journal_data(source, jcr)
            results.append(journal_data)
            fetched_this_run += 1
        else:
            not_found += 1

        # Save progress every 100 journals
        if (i + 1) % 100 == 0:
            save_json_atomic(OUTPUT_FILE, results)

        time.sleep(0.1)  # Rate limiting

    # Final save
    save_json_atomic(OUTPUT_FILE, results)

    print(f"\n--- Summary ---")
    print(f"Retained existing: {resume_stats['retained']}")
    print(f"Refreshed JCR metadata: {resume_stats['refreshed']}")
    print(f"Pruned outside current whitelist: {resume_stats['pruned']}")
    print(f"Missing before fetch: {resume_stats['missing']}")
    print(f"Fetched this run: {fetched_this_run}")
    print(f"Not found in OpenAlex: {not_found}")
    print(f"Total output: {len(results)}")

    # Subject distribution
    from collections import Counter
    subjects = Counter(j.get("_jcr_subject", "unknown") for j in results)
    print(f"\nTop 10 subjects in fetched data:")
    for subj, count in subjects.most_common(10):
        print(f"  {subj}: {count}")


if __name__ == "__main__":
    main()
