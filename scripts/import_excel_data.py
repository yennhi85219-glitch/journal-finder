"""
import_excel_data.py - 从 JCR 和中科院分区 Excel 导入数据到 manual_supplement.json

Use --jcr-file and --cas-file to provide the licensed Excel sources explicitly.
Without them, the script searches common user document directories.

匹配逻辑：通过 ISSN 与我们数据库中的期刊匹配
"""

import argparse
import json
import os
import re
import unicodedata
import openpyxl
from difflib import get_close_matches
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / "data"
RAW_DIR = DATA_DIR / "raw"
JCR_FILE = Path.home() / "Downloads/期刊/2026年度JCR期刊名单（完整版）.xlsx"
CAS_FILE = Path.home() / "Downloads/期刊/2025中科院分区表完整版（附2023vs2025对比版）.xlsx"
VALID_JCR_QUARTILES = {"Q1", "Q2", "Q3", "Q4"}
REVIEW_TYPE_PROVENANCE_FIELDS = {
    "review_type_source",
    "review_type_source_url",
    "review_type_last_checked",
    "review_type_verified",
}


def find_input_file(default_path, patterns):
    """Find a local Excel source even if the Downloads folder was reorganized."""
    if default_path.exists():
        return default_path

    roots = [
        Path.home() / "Downloads",
        Path.home() / "Desktop",
        Path.home() / "Documents",
    ]
    for root in roots:
        if not root.exists():
            continue
        for pattern in patterns:
            matches = sorted(root.rglob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
            if matches:
                return matches[0]
    return default_path


def write_json_atomic(path, data):
    """Write JSON beside its destination, then atomically replace the output."""
    tmp_path = path.with_name(f".{path.name}.tmp")
    try:
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            f.flush()
            os.fsync(f.fileno())
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def write_text_atomic(path, text):
    """Atomically replace a generated text report."""
    tmp_path = path.with_name(f".{path.name}.tmp")
    try:
        with open(tmp_path, "w", encoding="utf-8") as f:
            f.write(text)
            f.flush()
            os.fsync(f.fileno())
        tmp_path.replace(path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def normalize_jcr_quartile(value):
    """Convert missing JCR sentinels to None."""
    if value is None:
        return None
    normalized = str(value).strip().upper()
    if normalized in {"", "N/A", "NA", "-", "—"}:
        return None
    return normalized


def verified_review_type(entry):
    """Keep review type only when its source or verification is recorded."""
    value = entry.get("review_type")
    if not value:
        return None
    if not any(entry.get(field) for field in REVIEW_TYPE_PROVENANCE_FIELDS):
        return None
    normalized = str(value).strip().lower().replace("-", "_").replace(" ", "_")
    if normalized not in {"single_blind", "double_blind"}:
        return None
    return normalized


def normalize_supplement_entry(entry):
    """Migrate legacy placeholder values and fill stable nullable fields."""
    entry["jcr_quartile"] = normalize_jcr_quartile(entry.get("jcr_quartile"))
    if (
        entry["jcr_quartile"] is not None
        and entry["jcr_quartile"] not in VALID_JCR_QUARTILES
    ):
        raise ValueError(f"Invalid JCR quartile: {entry['jcr_quartile']!r}")
    entry["review_type"] = verified_review_type(entry)
    entry.setdefault("cas_zone", None)
    entry.setdefault("impact_factor", None)
    entry.setdefault("apc_usd", None)
    entry.setdefault("apc_waiver", None)
    entry.setdefault("oa_type", None)
    entry.setdefault("word_limit_min", None)
    entry.setdefault("word_limit_max", None)
    entry.setdefault("warning_tags", [])
    entry.setdefault("notes", "")
    entry.setdefault("last_verified", "2026-07")
    return entry


def clean_issn(issn):
    if not issn:
        return None
    issn = str(issn).upper().strip()
    issn = re.sub(r"[^0-9X]", "", issn)
    if len(issn) != 8:
        return None
    return f"{issn[:4]}-{issn[4:]}"


def normalize_title(title):
    """Normalize journal titles for robust cross-source matching."""
    if not title:
        return None
    text = unicodedata.normalize("NFKC", str(title)).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    words = [w for w in text.split() if w not in {"the", "a", "an", "journal"}]
    normalized = " ".join(words).strip()
    return normalized or None


def title_variants(title):
    norm = normalize_title(title)
    if not norm:
        return set()
    variants = {norm}
    replacements = [
        (" and ", " "),
        (" international ", " "),
        (" review ", " "),
    ]
    for old, new in replacements:
        if old in f" {norm} ":
            variants.add(normalize_title(norm.replace(old.strip(), new.strip())))
    return {v for v in variants if v}


def load_our_journals():
    """Load all ISSNs and names from final and raw databases."""
    issn_to_journal = {}  # issn -> issn_l mapping
    issn_l_to_name = {}
    name_to_issn_l = {}
    ambiguous_names = set()

    def add_name(name, issn_l):
        for variant in title_variants(name):
            existing = name_to_issn_l.get(variant)
            if existing and existing != issn_l:
                ambiguous_names.add(variant)
                continue
            name_to_issn_l[variant] = issn_l

    def add_issn(issn, issn_l):
        cleaned = clean_issn(issn)
        if cleaned:
            issn_to_journal[cleaned] = issn_l
            issn_to_journal[cleaned.replace("-", "")] = issn_l

    for filename in ["journals_ssci.json", "journals_economics.json", "journals_demography.json"]:
        filepath = DATA_DIR / filename
        if not filepath.exists():
            continue
        with open(filepath, "r", encoding="utf-8") as f:
            journals = json.load(f)
            for j in journals:
                issn_l = j.get("issn_l")
                if not issn_l:
                    continue
                issn_l_to_name[issn_l] = j.get("name", "")
                add_name(j.get("name"), issn_l)
                add_name(j.get("abbreviation"), issn_l)
                add_issn(issn_l, issn_l)
                all_issns = j.get("issn", []) if isinstance(j.get("issn"), list) else []
                for issn in all_issns:
                    add_issn(issn, issn_l)

    for filename in ["sources_ssci_all.json", "sources_economics.json", "sources_demography.json"]:
        filepath = RAW_DIR / filename
        if not filepath.exists():
            continue
        with open(filepath, "r", encoding="utf-8") as f:
            sources = json.load(f)
            for source in sources:
                issn_l = source.get("issn_l")
                if not issn_l:
                    continue
                issn_l_to_name.setdefault(issn_l, source.get("name", ""))
                add_name(source.get("name"), issn_l)
                for alt in source.get("alternate_titles", []) or []:
                    add_name(alt, issn_l)
                for issn in source.get("issn", []) or []:
                    add_issn(issn, issn_l)
                add_issn(issn_l, issn_l)

    for name in ambiguous_names:
        name_to_issn_l.pop(name, None)

    return issn_to_journal, issn_l_to_name, name_to_issn_l


def match_by_name(name, name_to_issn_l):
    for variant in title_variants(name):
        issn_l = name_to_issn_l.get(variant)
        if issn_l:
            return issn_l, "name_exact"
    return None, None


def build_fuzzy_index(name_to_issn_l):
    """Bucket normalized names so fuzzy fallback stays fast and conservative."""
    buckets = {}
    for name in name_to_issn_l:
        tokens = name.split()
        if not tokens:
            continue
        buckets.setdefault(tokens[0], []).append(name)
    return buckets


def match_by_fuzzy_name(name, name_to_issn_l, fuzzy_index=None, cutoff=0.94):
    norm = normalize_title(name)
    if not norm:
        return None, None
    tokens = norm.split()
    if fuzzy_index and tokens:
        candidates = fuzzy_index.get(tokens[0], [])
        lower_len = max(1, int(len(norm) * 0.75))
        upper_len = int(len(norm) * 1.25) + 1
        candidates = [c for c in candidates if lower_len <= len(c) <= upper_len]
    else:
        candidates = list(name_to_issn_l.keys())
    if not candidates:
        return None, None
    matches = get_close_matches(norm, candidates, n=1, cutoff=cutoff)
    if not matches:
        return None, None
    return name_to_issn_l[matches[0]], f"name_fuzzy:{matches[0]}"


def coverage_summary(supplement):
    """Summarize decision-field coverage in manual_supplement.json."""
    total = len(supplement)
    if not total:
        return {}
    fields = ["jcr_quartile", "cas_zone", "impact_factor"]
    return {
        field: sum(1 for entry in supplement.values() if entry.get(field) is not None)
        for field in fields
    }


def import_jcr():
    """Import JCR data - extract ISSN, IF, JIF quartile, publisher, subject category."""
    print("Loading JCR 2026 data...")
    wb = openpyxl.load_workbook(JCR_FILE, read_only=True)
    ws = wb[wb.sheetnames[0]]

    jcr_data = {}  # issn -> record

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: 0=rank, 1=name, 2=abbr, 3=ISSN, 4=eISSN, 5=publisher, 6=subject,
        # 7=JIF, 8=JIF_quartile, ...
        if not row or not row[3]:
            continue

        issn = str(row[3]).strip().upper() if row[3] else None
        eissn = str(row[4]).strip().upper() if row[4] else None
        name = str(row[1]).strip() if row[1] else ""

        record = {
            "name": name,
            "abbreviation": str(row[2]).strip() if row[2] else None,
            "publisher": str(row[5]).strip() if row[5] else None,
            "subject_category": str(row[6]).strip() if row[6] else None,
            "impact_factor": float(row[7]) if row[7] and row[7] != "N/A" else None,
            "jcr_quartile": normalize_jcr_quartile(row[8]),
            "jif_percentile": float(row[9]) if row[9] else None,
            "five_year_if": float(row[15]) if row[15] else None,
            "gold_oa_pct": float(row[29]) if row[29] else None,
            "subject_detail": str(row[31]).strip() if row[31] else None,
        }

        if issn:
            jcr_data[issn] = record
        if eissn:
            jcr_data[eissn] = record

    wb.close()
    print(f"  Loaded {len(jcr_data)} ISSN entries from JCR")
    return jcr_data


def import_cas():
    """Import CAS zoning data - match by journal name."""
    print("Loading CAS 2025 zoning data...")
    wb = openpyxl.load_workbook(CAS_FILE, read_only=True)

    # Use '完整版' sheet
    ws = wb['完整版']
    # Headers: 期刊名称, 2025分区, Top, Open Access

    cas_data = {}  # upper name -> record

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip().upper()
        cas_data[name] = {
            "cas_zone": int(row[1]) if row[1] else None,
            "is_top": str(row[2]).strip() == "是" if row[2] else False,
            "is_oa": str(row[3]).strip() == "是" if row[3] else False,
        }

    wb.close()

    # Also load comparison sheet for more data
    wb2 = openpyxl.load_workbook(CAS_FILE, read_only=True)
    ws2 = wb2['2023 vs 2025分区对比']
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip().upper()
        if name not in cas_data:
            cas_data[name] = {
                "cas_zone": int(row[1]) if row[1] else None,
                "is_top": str(row[3]).strip() == "是" if row[3] else False,
                "is_oa": str(row[4]).strip() == "是" if row[4] else False,
            }
    wb2.close()

    print(f"  Loaded {len(cas_data)} journal entries from CAS")
    return cas_data


def main():
    global JCR_FILE, CAS_FILE

    parser = argparse.ArgumentParser(description="Import JCR and CAS Excel data")
    parser.add_argument("--jcr-file", type=Path)
    parser.add_argument("--cas-file", type=Path)
    args = parser.parse_args()

    JCR_FILE = (
        args.jcr_file.expanduser().resolve()
        if args.jcr_file
        else find_input_file(
            JCR_FILE,
            ["*JCR*.xlsx", "*jcr*.xlsx", "*期刊名单*.xlsx"],
        )
    )
    CAS_FILE = (
        args.cas_file.expanduser().resolve()
        if args.cas_file
        else find_input_file(
            CAS_FILE,
            ["*中科院*.xlsx", "*分区*.xlsx"],
        )
    )
    print(f"JCR source: {JCR_FILE}")
    print(f"CAS source: {CAS_FILE}")

    # Load our journal database ISSNs and names
    issn_to_issn_l, issn_l_to_name, name_to_issn_l = load_our_journals()
    fuzzy_index = build_fuzzy_index(name_to_issn_l)
    print(
        f"Our database: {len(issn_l_to_name)} journals, "
        f"{len(issn_to_issn_l)} ISSNs indexed, {len(name_to_issn_l)} names indexed"
    )

    # Load existing supplement
    supplement_path = DATA_DIR / "manual_supplement.json"
    if supplement_path.exists():
        with open(supplement_path, "r", encoding="utf-8") as f:
            supplement = json.load(f)
    else:
        supplement = {}
    if not isinstance(supplement, dict):
        raise ValueError(f"{supplement_path} must contain a JSON object")
    for entry in supplement.values():
        normalize_supplement_entry(entry)
    print(f"Existing supplement: {len(supplement)} entries")
    before_coverage = coverage_summary(supplement)

    # Import JCR
    jcr_data = import_jcr()

    # Import CAS
    cas_data = import_cas()

    # Match JCR data to our journals: ISSN first, then conservative title matching.
    jcr_matched = {}
    jcr_unmatched = []
    jcr_methods = {"issn": 0, "name_exact": 0, "name_fuzzy": 0}
    for issn, jcr_record in jcr_data.items():
        cleaned = clean_issn(issn)
        issn_clean = cleaned.replace("-", "") if cleaned else str(issn).replace("-", "").strip()
        issn_with_dash = cleaned or str(issn).strip()

        # Try to find this ISSN in our database
        issn_l = issn_to_issn_l.get(issn_with_dash)
        if not issn_l:
            issn_l = issn_to_issn_l.get(issn_clean)
        match_method = "issn" if issn_l else None
        if not issn_l:
            issn_l, match_method = match_by_name(jcr_record.get("name"), name_to_issn_l)
        if not issn_l:
            issn_l, match_method = match_by_fuzzy_name(
                jcr_record.get("name"),
                name_to_issn_l,
                fuzzy_index=fuzzy_index,
                cutoff=0.96,
            )
        if not issn_l:
            jcr_unmatched.append(jcr_record.get("name") or str(issn))
            continue
        if issn_l in jcr_matched:
            continue

        # Found a match - update supplement
        if issn_l not in supplement:
            supplement[issn_l] = {}

        entry = supplement[issn_l]
        if jcr_record.get("impact_factor") is not None:
            entry["impact_factor"] = jcr_record["impact_factor"]
        entry["jcr_quartile"] = normalize_jcr_quartile(jcr_record.get("jcr_quartile"))
        if jcr_record.get("subject_category"):
            entry["subject_category"] = jcr_record["subject_category"]
        if jcr_record.get("subject_detail"):
            entry["subject_detail"] = jcr_record["subject_detail"]
        if jcr_record.get("five_year_if") is not None:
            entry["five_year_if"] = jcr_record["five_year_if"]
        if jcr_record.get("gold_oa_pct") is not None:
            entry["gold_oa_pct"] = jcr_record["gold_oa_pct"]

        entry["jcr_match_method"] = match_method
        entry["jcr_source_name"] = jcr_record.get("name")
        entry["last_verified"] = "2026-07"
        jcr_matched[issn_l] = True
        if match_method == "issn":
            jcr_methods["issn"] += 1
        elif match_method == "name_exact":
            jcr_methods["name_exact"] += 1
        elif match_method and match_method.startswith("name_fuzzy"):
            jcr_methods["name_fuzzy"] += 1

    print(f"\nJCR matched: {len(jcr_matched)} journals")
    print(f"  by method: {jcr_methods}")

    # Match CAS data to our journals (by name)
    cas_matched = {}
    cas_unmatched = []
    cas_methods = {"name_exact": 0, "name_fuzzy": 0}
    for cas_name, cas_record in cas_data.items():
        issn_l, match_method = match_by_name(cas_name, name_to_issn_l)
        if not issn_l:
            issn_l, match_method = match_by_fuzzy_name(
                cas_name,
                name_to_issn_l,
                fuzzy_index=fuzzy_index,
                cutoff=0.96,
            )
        if not issn_l:
            cas_unmatched.append(cas_name)
            continue
        if issn_l in cas_matched:
            continue

        if issn_l not in supplement:
            supplement[issn_l] = {}

        entry = supplement[issn_l]
        if cas_record.get("cas_zone") is not None:
            entry["cas_zone"] = cas_record["cas_zone"]
        if cas_record.get("is_top"):
            entry["cas_top"] = True
        entry["cas_match_method"] = match_method
        entry["cas_source_name"] = cas_name
        cas_matched[issn_l] = True
        if match_method == "name_exact":
            cas_methods["name_exact"] += 1
        elif match_method and match_method.startswith("name_fuzzy"):
            cas_methods["name_fuzzy"] += 1

    print(f"CAS matched: {len(cas_matched)} journals")
    print(f"  by method: {cas_methods}")

    # Ensure all entries have required fields
    for entry in supplement.values():
        normalize_supplement_entry(entry)

    # Save
    write_json_atomic(supplement_path, supplement)

    print(f"\nFinal supplement: {len(supplement)} entries saved to {supplement_path}")

    after_coverage = coverage_summary(supplement)
    print("\n--- Coverage summary ---")
    total = len(supplement) or 1
    for field, count in after_coverage.items():
        before = before_coverage.get(field, 0)
        print(f"  {field}: {count}/{len(supplement)} ({count / total:.1%}), +{count - before}")

    write_text_atomic(
        DATA_DIR / "unmatched_jcr_names.txt",
        "\n".join(sorted(set(jcr_unmatched))) + "\n",
    )
    write_text_atomic(
        DATA_DIR / "unmatched_cas_names.txt",
        "\n".join(sorted(set(cas_unmatched))) + "\n",
    )
    print("  Wrote unmatched lists: unmatched_jcr_names.txt, unmatched_cas_names.txt")

    # Show some examples
    print("\n--- Sample entries ---")
    sample_issns = ["0002-8282", "0033-5533", "0070-3370", "0933-1433"]
    for issn_l in sample_issns:
        if issn_l in supplement:
            e = supplement[issn_l]
            name = issn_l_to_name.get(issn_l, "?")
            print(f"  {name}: JCR={e.get('jcr_quartile')}, CAS={e.get('cas_zone')}, IF={e.get('impact_factor')}")


if __name__ == "__main__":
    main()
